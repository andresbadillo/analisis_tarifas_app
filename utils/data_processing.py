"""Módulo para el procesamiento de datos de tarifas."""

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
import io
from typing import Optional, Dict, Any

def cargar_tabla_desde_excel(file_content: Any) -> Optional[pd.DataFrame]:
    """
    Carga específicamente la tabla 'TablaDatos' desde la hoja 'Hojadedatos' del archivo Excel.
    
    Args:
        file_content: Contenido del archivo Excel cargado.
    
    Returns:
        DataFrame con los datos cargados o None si hay error.
    """
    try:
        with io.BytesIO(file_content.read()) as file:
            wb = load_workbook(file, data_only=True)
            
            if "Hojadedatos" not in wb.sheetnames:
                st.error("❌ No se encontró la hoja 'Hojadedatos' en el archivo")
                return None
            
            sheet = wb["Hojadedatos"]
            
            if "TablaDatos" not in sheet.tables:
                st.error("❌ No se encontró la tabla 'TablaDatos' en la hoja 'Hojadedatos'")
                return None
            
            table = sheet.tables["TablaDatos"]
            table_range = table.ref
            data = sheet[table_range]
            rows = [[cell.value for cell in row] for row in data]
            
            if len(rows) < 2:
                st.error("❌ La tabla 'TablaDatos' está vacía")
                return None
            
            df = pd.DataFrame(rows[1:], columns=rows[0])
            st.success("✅ Tabla 'TablaDatos' encontrada y cargada desde la hoja 'Hojadedatos'")
            
            return df
            
    except Exception as e:
        st.error(f"❌ Error al cargar la tabla: {str(e)}")
        return None

def procesar_df_tarifas(df: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Procesa el DataFrame de tarifas para obtener las columnas necesarias y realizar las transformaciones.
    
    Args:
        df: DataFrame con los datos crudos.
    
    Returns:
        DataFrame procesado o None si hay error.
    """
    try:
        columnas_necesarias = {'MERCADO', 'FECHA', 'COMERCIALIZADOR', 'NT', 'G', 'C', 'CU'}
        if not columnas_necesarias.issubset(df.columns):
            st.error("❌ La tabla no contiene todas las columnas necesarias.")
            st.info("ℹ️ Las columnas requeridas son: MERCADO, FECHA, COMERCIALIZADOR, NT, G, C, CU")
            return None

        df['FECHA'] = pd.to_datetime(df['FECHA'])

        columnas_numericas = ['G', 'C', 'CU']
        for col in columnas_numericas:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        columnas_orden = ['FECHA', 'MERCADO', 'COMERCIALIZADOR', 'NT', 'G', 'C', 'CU']
        df = df[columnas_orden]
        df['FECHA'] = df['FECHA'].dt.strftime('%Y-%m')

        return df

    except Exception as e:
        st.error(f"❌ Error al procesar los datos: {str(e)}")
        return None 